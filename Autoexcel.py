import os
import sys
import openpyxl
from openpyxl.styles import NamedStyle
wbname = sys.argv[-1]
wb = openpyxl.load_workbook(wbname)
print("Excel processing started for :", wbname)
index1 = wb["Index"]
timeser = wb["Timeseries"]
ws1 = wb.create_sheet("Sheet_A")
ws1.title = "New Timeseries"
result = wb[ws1.title]
#Date style for new sheet "MMM-YY" format
nsmmmyy=NamedStyle(name='cd1',number_format="MMM-YY")
for x in range(1, timeser.max_row+1):
    result.cell(row=x, column=1).value = timeser.cell(row=x, column=1).value
    if x == 1:
        continue
    else:
        try:
            result.cell(row=x, column=1).style = nsmmmyy
        except ValueError:
            result.cell(row=x, column=1).style = 'cd1'
i = 2
w = i
#Loop thorough WLCODE in 1st row of Time series sheet till last column(spaces) is reached
while timeser.cell(row=1, column=i).value is not None:
    matchfound = "n"
    #Loop to find the elevation from index sheet
    for x in range(2, index1.max_row + 1):
        if index1.cell(row=x, column=9).value == timeser.cell(row=1, column=i).value:
            elevation = index1.cell(row=x, column=10).value
            if elevation == None or elevation == "" or elevation == "" or elevation == 0:
                matchfound = "x"
            else:
                matchfound = "y"
    else:
        #When elevation is found from index sheet
        if matchfound == "y":
            #Move WLCODE header to result sheet.
            result.cell(row=1, column=w).value = timeser.cell(row=1, column=i).value
            for j in range(2, timeser.max_row + 1):
                tele = timeser.cell(row=j, column=i).value
                #If timeseries value is NULL move NULL to result else result = elevation - timeseries
                if tele == None or tele == "" or tele == " " or tele == "0" or tele == 0  :
                    if tele == 0:
                        result.cell(row=j, column=w).value = None
                    else:
                        result.cell(row=j, column=w).value = timeser.cell(row=j, column=i).value
                else:
                    result1 = elevation - tele
                    result.cell(row=j, column=w).value = result1
            w=w+1
        #If elevation has no value print "elevation not found"
        elif matchfound == "x":
            print("Elevation invalid for WLCODE :", timeser.cell(row=1, column=i).value)
        # If no row for WLCODE in index sheet
        else:
            print("Elevation not found for WLCODE :", timeser.cell(row=1, column=i).value)
    i = i + 1
wb.save(wbname)
print("Excel processing completed for :" , wbname)